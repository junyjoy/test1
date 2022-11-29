import openpyxl
import random

path = 'weight_height.xlsx'

wb = openpyxl.load_workbook(path)
ws = wb['Sheet1']

#print(ws.cell(row=2,column=2).value)

x_wt = []
y_ht = []

# 1. 성별 2. 키  3. 몸무게

max_row = ws.max_row

for i in range(2, max_row+1):
    ht = ws.cell(row=i, column=2).value
    wt = ws.cell(row=i, column=3).value
    x_wt.append(wt)
    y_ht.append(ht)
#print(x_wt, y_ht)

batch_size = 64
step_per_epoch = int(len(x_wt)/batch_size)
epochs = 150

w_n = 0.001
a = 0.0001
b_n = 0.0000001
min_loss = 500

for epoch in range(1, epochs+1):
    random.Random(123).shuffle(x_wt)
    random.Random(123).shuffle(y_ht)

    for step in range(0,step_per_epoch):
        begin = step * batch_size
        end = begin + batch_size

        x_batch = x_wt[begin:end]
        y_batch = y_ht[begin:end]

        n = batch_size

        sigma = 0
        for i in range(n):
            sigma += (((w_n*x_batch[i])-y_batch[i]+b_n)*x_batch[i])
        w_s = sigma * 2/n 

        sigma = 0
        for i in range(n):
            sigma += (((w_n*x_batch[i])-y_batch[i]+b_n)*x_batch[i])
        b_s = sigma * 2/n 


        w_n1 = w_n - (a * w_s)
        b_n1 = b_n - (a*b_s)
        sigma = 0
        for i in range(n):
            sigma += ((w_n*x_batch[i])-y_batch[i])**2
            loss = sigma/n
            w_n = w_n1

            if loss < min_loss:
                best_w = w_n1
                best_b = b_n1
                min_loss = loss



        print('Epoch=', epoch, '가중치=', w_n1,'bias=', b_n1, 'Loss=', loss)
print(min_loss, best_w, best_b)

my_weight = 120
my_height = (my_weight * best_w) + best_b

print(my_height)